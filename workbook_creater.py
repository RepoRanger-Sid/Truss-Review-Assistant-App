import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
import os
import re

def create_workbook(pdf_file_name, truss_dictionary):
    # CREATING A WORKBOOK AND SAVING IT IN THE CURRENT WORKING DIRECTORY
    wb = openpyxl.Workbook()
    ws = wb.active
    file_name = 'Truss Review Results_' + pdf_file_name
    ws.title = "Summary"
    directory_path = os.getcwd()
    file_path = directory_path + '/' + file_name + '.xlsx'

    # Dictionary with keys from 'A' to 'T' and their respective values
    cell_values = {
        'A': {'A1': 'TRUSS LABEL'},
        'B': {'B1': 'TOP CHORD DEAD LOAD', 'B2': '(PSF)'},
        'C': {'C1': 'BOTTOM CHORD DEAD LOAD', 'C2': '(PSF)'},
        'D': {'D1': 'TOP CHORD LIVE LOAD', 'D2': '(PSF)'},
        'E': {'E1': 'BOTTOM CHORD LIVE LOAD', 'E2': '(PSF)'},
        'F': {'F1': 'SLOPE', 'F2': '(/12)'},
        'G': {'G1': 'TRUSS SPACING', 'G2': '(FT)'},
        'H': {'H1': 'WIND STANDARD'},
        'I': {'I1': 'WIND SPEED', 'I2': '(MPH)'},
        'J': {'J1': 'RISK CATEGORY'},
        'K': {'K1': 'BUILDING CODE'},
        'L': {'L1': 'MAX TOP CHORD STRESS'},
        'M': {'M1': 'MAX BOTTOM CHORD STRESS'},
        'N': {'N1': 'MAX WEB CHORD STRESS'},
        'O': {'O1': 'LIVE LOAD DEFLECTION CRITERIA', 'O2': '(L/)'},
        'P': {'P1': 'TOTAL LOAD DEFLECTION CRITERIA', 'P2': '(L/)'},
        'Q': {'Q1': 'CALCULATED LIVE LOAD DEFLECTION', 'Q2': '(L/)'},
        'R': {'R1': 'CALCULATED TOTAL LOAD DEFLECTION', 'R2': '(L/)'},
        'S': {'S1': 'DRAG LOAD', 'S2': '(LB)'},
        'T': {'T1': 'WARNING'}
    }

    # Initializing the title row and unit row using the dictionary
    for column, values in cell_values.items():
        for cell, value in values.items():
            ws[cell] = value

    # SETTING UP ROW AND COLUMN SIZE
    ws.row_dimensions[1].height = 70
    ws.row_dimensions[2].height = 15
    ws.freeze_panes = 'B3'

    for column in cell_values.keys():
        # SETTING CELL WIDTH
        ws.column_dimensions[column].width = 20

        # SETTING FONT ALIGNMENT AND FONT STYLE
        ws[f'{column}1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws[f'{column}2'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'{column}1'].font = Font(bold=True)
        ws[f'{column}2'].font = Font(bold=True)

        # SETTING CELL BORDER STYLE
        thin_border = Side(border_style='thin', color='000000')
        border_format = Border(bottom=thin_border)
        ws[f'{column}2'].border = border_format

    # SETTING UP CELL BORDER FOR FIRST COLUMN ONLY : AT TRUSS LABELS
    for i in range(len(truss_dictionary.keys())):
        thin_border = Side(border_style='thin', color='000000')
        border_format = Border(right=thin_border)
        ws[f'A{i + 4}'].border = border_format

    # Save the workbook
    wb.save(file_path)
    return file_path


# UNPACKING DICTONARY AND STORING VALUES IN CELLS
def unpack_and_store_values(file_path, truss_dictionary):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    type_non_float = re.compile(r'[*a-zA-Z-]+')
    row_num = 4  # Initialize row_num to start from the fourth row

    for truss, values in truss_dictionary.items():
        for col_num, value in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col_num)
            if re.search(type_non_float, value):
                cell.value = value
            else:
                cell.value = float(value)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        row_num += 1

    wb.save(file_path)