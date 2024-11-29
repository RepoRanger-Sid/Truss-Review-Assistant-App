import openpyxl
from openpyxl.styles import Font, PatternFill

def format_cells(sheet, cells, font_color='00FF0000', fill_color='00FFCC99'):
    """Format Excel cells with specified colors"""
    for cell in cells:
        if font_color is None and fill_color is None:
            sheet[cell].font = Font()
            sheet[cell].fill = PatternFill(fill_type=None)
        else:
            sheet[cell].font = Font(color=font_color)
            sheet[cell].fill = PatternFill(
                start_color=fill_color,
                end_color=fill_color,
                fill_type='solid'
            )

class DataValidator:
    def __init__(self, workbook_path, input_data):
        self.workbook_path = workbook_path
        self.input_data = input_data
        self.wb = openpyxl.load_workbook(workbook_path)
        self.ws = self.wb.active

    def save_workbook(self):
        """Save workbook after modifications"""
        self.wb.save(self.workbook_path)

    def is_float(self, value):
        """Check if a value can be converted to float"""
        try:
            float(value)
            return True
        except ValueError:
            return False

    def check_dead_load(self):
        """Check dead load values"""
        for row in range(4, self.ws.max_row + 1):
            try:
                b_value = self.ws[f'B{row}'].value or 0
                c_value = self.ws[f'C{row}'].value or 0
                f_value = self.ws[f'F{row}'].value or 0

                if not self.is_float(b_value) or not self.is_float(c_value) or not self.is_float(f_value):
                    continue

                b_value = float(b_value)
                c_value = float(c_value)
                f_value = float(f_value)
                
                sum_bc = b_value + c_value
                cells = [f'B{row}', f'C{row}']

                if f_value != 0:  # Roof
                    if sum_bc >= float(self.input_data['design_roof_dead_load']):
                        format_cells(self.ws, cells, font_color=None, fill_color=None)
                    else:
                        format_cells(self.ws, cells, font_color='00FF0000', fill_color='00FFCC99')
                else:  # Floor
                    if sum_bc >= float(self.input_data['design_floor_dead_load']):
                        format_cells(self.ws, cells, font_color=None, fill_color=None)
                    else:
                        format_cells(self.ws, cells, font_color='00FF0000', fill_color='00FFCC99')
            except (ValueError, TypeError) as e:
                print(f"Error processing row {row}: {str(e)}")

    def check_live_load(self):
        """Check live load values"""
        for row in range(4, self.ws.max_row + 1):
            try:
                d_value = self.ws[f'D{row}'].value or 0
                e_value = self.ws[f'E{row}'].value or 0
                f_value = self.ws[f'F{row}'].value or 0

                if not self.is_float(d_value) or not self.is_float(e_value) or not self.is_float(f_value):
                    continue

                d_value = float(d_value)
                e_value = float(e_value)
                f_value = float(f_value)
                
                sum_de = d_value + e_value
                cells = [f'D{row}', f'E{row}']

                if f_value >= 4:  # More slope
                    if sum_de >= float(self.input_data['design_roof_live_load_more_slope']):
                        format_cells(self.ws, cells, font_color=None, fill_color=None)
                    else:
                        format_cells(self.ws, cells, font_color='00FF0000', fill_color='00FFCC99')
                elif f_value != 0:  # Less slope
                    if sum_de >= float(self.input_data['design_roof_live_load_less_slope']):
                        format_cells(self.ws, cells, font_color=None, fill_color=None)
                    else:
                        format_cells(self.ws, cells, font_color='00FF0000', fill_color='00FFCC99')
                else:  # Floor
                    if sum_de >= float(self.input_data['design_floor_live_load']):
                        format_cells(self.ws, cells, font_color=None, fill_color=None)
                    else:
                        format_cells(self.ws, cells, font_color='00FF0000', fill_color='00FFCC99')
            except (ValueError, TypeError) as e:
                print(f"Error processing row {row}: {str(e)}")

    def check_truss_spacing(self):
        """Check truss spacing"""
        for row in range(4, self.ws.max_row + 1):
            try:
                g_value = self.ws[f'G{row}'].value or 0

                if not self.is_float(g_value):
                    continue

                g_value = float(g_value)
                if g_value <= float(self.input_data['actual_truss_spacing']):
                    format_cells(self.ws, [f'G{row}'], font_color=None, fill_color=None)
                else:
                    format_cells(self.ws, [f'G{row}'], font_color='00FF0000', fill_color='00FFCC99')
            except (ValueError, TypeError) as e:
                print(f"Error processing row {row}: {str(e)}")

    def check_codes_and_standards(self):
        """Check building codes and standards"""
        for row in range(4, self.ws.max_row + 1):
            try:
                f_value = self.ws[f'F{row}'].value or 0
                h_value = str(self.ws[f'H{row}'].value or '')
                i_value = self.ws[f'I{row}'].value or 0
                j_value = str(self.ws[f'J{row}'].value or '')
                k_value = str(self.ws[f'K{row}'].value or '')

                if not self.is_float(f_value) or not self.is_float(i_value):
                    continue

                f_value = float(f_value)
                i_value = float(i_value)

                if f_value != 0:
                    if h_value == self.input_data['design_wind_standard']:
                        format_cells(self.ws, [f'H{row}'], font_color=None, fill_color=None)
                    else:
                        format_cells(self.ws, [f'H{row}'], font_color='00FF0000', fill_color='00FFCC99')
                
                if i_value >= float(self.input_data['design_wind_speed']):
                    format_cells(self.ws, [f'I{row}'], font_color=None, fill_color=None)
                else:
                    format_cells(self.ws, [f'I{row}'], font_color='00FF0000', fill_color='00FFCC99')
                
                if j_value == self.input_data['design_risk_category']:
                    format_cells(self.ws, [f'J{row}'], font_color=None, fill_color=None)
                else:
                    format_cells(self.ws, [f'J{row}'], font_color='00FF0000', fill_color='00FFCC99')
                
                if k_value == self.input_data['design_building_code']:
                    format_cells(self.ws, [f'K{row}'], font_color=None, fill_color=None)
                else:
                    format_cells(self.ws, [f'K{row}'], font_color='00FF0000', fill_color='00FFCC99')
            except (ValueError, TypeError) as e:
                print(f"Error processing row {row}: {str(e)}")

    def check_member_stresses(self):
        """Check member stress values"""
        for row in range(4, self.ws.max_row + 1):
            try:
                for col in ['L', 'M', 'N']:
                    value = self.ws[f'{col}{row}'].value or 0

                    if not self.is_float(value):
                        continue

                    value = float(value)
                    if value < 1:
                        format_cells(self.ws, [f'{col}{row}'], font_color=None, fill_color=None)
                    else:
                        format_cells(self.ws, [f'{col}{row}'], font_color='00FF0000', fill_color='00FFCC99')
            except (ValueError, TypeError) as e:
                print(f"Error processing row {row}: {str(e)}")

    def check_deflections(self):
        """Check deflection values"""
        for row in range(4, self.ws.max_row + 1):
            try:
                f_value = self.ws[f'F{row}'].value or 0
                o_value = self.ws[f'O{row}'].value or 0
                p_value = self.ws[f'P{row}'].value or 0
                q_value = self.ws[f'Q{row}'].value
                r_value = self.ws[f'R{row}'].value

                if not self.is_float(f_value) or not self.is_float(o_value) or not self.is_float(p_value):
                    continue

                f_value = float(f_value)
                o_value = float(o_value)
                p_value = float(p_value)

                if f_value != 0:  # Roof
                    if o_value >= float(self.input_data['design_roof_live_load_deflection_criteria']):
                        format_cells(self.ws, [f'O{row}'], font_color=None, fill_color=None)
                    else:
                        format_cells(self.ws, [f'O{row}'], font_color='00FF0000', fill_color='00FFCC99')
                    
                    if p_value >= float(self.input_data['design_roof_total_load_deflection_criteria']):
                        format_cells(self.ws, [f'P{row}'], font_color=None, fill_color=None)
                    else:
                        format_cells(self.ws, [f'P{row}'], font_color='00FF0000', fill_color='00FFCC99')
                else:  # Floor
                    if o_value >= float(self.input_data['design_floor_live_load_deflection_criteria']):
                        format_cells(self.ws, [f'O{row}'], font_color=None, fill_color=None)
                    else:
                        format_cells(self.ws, [f'O{row}'], font_color='00FF0000', fill_color='00FFCC99')
                    
                    if p_value >= float(self.input_data['design_floor_total_load_deflection_criteria']):
                        format_cells(self.ws, [f'P{row}'], font_color=None, fill_color=None)
                    else:
                        format_cells(self.ws, [f'P{row}'], font_color='00FF0000', fill_color='00FFCC99')

                if isinstance(q_value, str) and q_value.lower() in ['n/r','n/a','****']:
                    format_cells(self.ws, [f'Q{row}'], font_color='00FF0000', fill_color='00FFCC99')
                else:
                    q_value = float(q_value or 0)
                    if f_value != 0:  # Roof
                        if q_value >= float(self.input_data['design_roof_live_load_deflection_criteria']):
                            format_cells(self.ws, [f'Q{row}'], font_color=None, fill_color=None)
                        else:
                            format_cells(self.ws, [f'Q{row}'], font_color='00FF0000', fill_color='00FFCC99')
                    else:  # Floor
                        if q_value >= float(self.input_data['design_floor_live_load_deflection_criteria']):
                            format_cells(self.ws, [f'Q{row}'], font_color=None, fill_color=None)
                        else:
                            format_cells(self.ws, [f'Q{row}'], font_color='00FF0000', fill_color='00FFCC99')

                if isinstance(r_value, str) and r_value.lower() in ['n/r','n/a','****']:
                    format_cells(self.ws, [f'R{row}'], font_color='00FF0000', fill_color='00FFCC99')
                else:
                    r_value = float(r_value or 0)
                    if f_value != 0:  # Roof
                        if r_value >= float(self.input_data['design_roof_total_load_deflection_criteria']):
                            format_cells(self.ws, [f'R{row}'], font_color=None, fill_color=None)
                        else:
                            format_cells(self.ws, [f'R{row}'], font_color='00FF0000', fill_color='00FFCC99')
                    else:  # Floor
                        if r_value >= float(self.input_data['design_floor_total_load_deflection_criteria']):
                            format_cells(self.ws, [f'R{row}'], font_color=None, fill_color=None)
                        else:
                            format_cells(self.ws, [f'R{row}'], font_color='00FF0000', fill_color='00FFCC99')

            except (ValueError, TypeError) as e:
                print(f"Error processing row {row}: {str(e)}")

    def check_warnings(self):
        """Check warning messages"""
        for row in range(4, self.ws.max_row + 1):
            if self.ws[f'T{row}'].value == 'WARNING':
                format_cells(self.ws, [f'T{row}'], font_color='00FF0000', fill_color='00FFCC99')
            else:
                format_cells(self.ws, [f'T{row}'], font_color=None, fill_color=None)

    def validate_all(self):
        """Run all validation checks"""
        try:
            self.check_dead_load()
            self.check_live_load()
            self.check_truss_spacing()
            self.check_codes_and_standards()
            self.check_member_stresses()
            self.check_deflections()
            self.check_warnings()
            self.save_workbook()
        except Exception as e:
            print(f"Error during validation: {str(e)}")

def assess(pdf_file_name,input_data):
    """Main execution function"""
    try:
        workbook_path = f'Truss Review Results_{pdf_file_name}.xlsx'
        
        # Get input criteria and validate data
        input_criteria = input_data
        validator = DataValidator(workbook_path, input_criteria)
        validator.validate_all()
        
        print("Data validation completed successfully")
    except Exception as e:
        print(f"Error in validation process: {str(e)}")
